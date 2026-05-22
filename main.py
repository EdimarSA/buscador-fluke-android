import json
import os
import re
from pathlib import Path

from kivy.app import App
from kivy.metrics import dp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.core.window import Window

from openpyxl import load_workbook

try:
    from plyer import filechooser
except Exception:
    filechooser = None

CONFIG_FILE = "config_android.json"
SHEET_NAME = "Sheet1"
HEADER_ROW = 5
DATA_START_ROW = 6

SEARCH_COLUMNS = ["Modelo", "Código", "EAN", "Descripción"]
DISPLAY_COLUMNS = ["Modelo", "Código", "EAN", "Descripción", "Dto", "PVP", "COSTE", "MARGEN", "PRECIO"]
MARGIN_OPTIONS = [0.10, 0.12, 0.15, 0.20, 0.33]
DEFAULT_MARGIN = 0.20

CANONICAL_COLUMNS = {
    "modelo": "Modelo",
    "codigo fluke": "Código",
    "código fluke": "Código",
    "codigofluke": "Código",
    "codigo": "Código",
    "código": "Código",
    "codigo ean": "EAN",
    "código ean": "EAN",
    "ean": "EAN",
    "descripcion": "Descripción",
    "descripción": "Descripción",
    "dto": "Dto",
    "descuento": "Dto",
    "pvp": "PVP",
}


def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip().lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ü": "u", "ñ": "n",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return re.sub(r"\s+", " ", text)


def to_number(value):
    if value is None:
        return 0.0

    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return 0.0

    text = text.replace("€", "").replace("%", "").replace(" ", "").replace("*", "")

    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except Exception:
        return 0.0


def is_numeric_discount(value):
    text = str(value).strip()

    if text == "":
        return True

    text = text.replace("€", "").replace("%", "").replace("*", "").replace(" ", "")

    if text == "":
        return False

    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        float(text)
        return True
    except Exception:
        return False


def parse_discount(value):
    text = str(value).strip()
    number = to_number(text)

    if "%" in text:
        return number / 100

    if number > 1:
        return number / 100

    return number


def format_money(value):
    try:
        return f"{float(value):,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "-" if value == "-" else ""


def format_percent(value):
    try:
        return f"{float(value) * 100:.1f}%"
    except Exception:
        return "-" if value == "-" else ""


def format_discount_for_display(value):
    try:
        if not is_numeric_discount(value):
            return str(value)

        text = str(value).strip()
        if text == "":
            return ""

        discount = parse_discount(text)
        suffix = "*" if "*" in text else ""
        return f"{discount * 100:.1f}%{suffix}"
    except Exception:
        return str(value)


def make_label(text, width=120, bold=False, color=(1, 1, 1, 1), bg=None):
    label = Label(
        text=str(text),
        size_hint_x=None,
        width=dp(width),
        size_hint_y=None,
        height=dp(42),
        halign="left",
        valign="middle",
        color=color,
        markup=bold,
    )
    label.bind(size=lambda instance, value: setattr(instance, "text_size", value))
    return label


class PriceSearchApp(App):
    def build(self):
        Window.clearcolor = (0.08, 0.08, 0.08, 1)

        self.excel_path = ""
        self.rows = []
        self.current_margin = DEFAULT_MARGIN
        self.result_rows = []

        self.root_layout = BoxLayout(orientation="vertical", padding=dp(8), spacing=dp(8))

        self.status_label = Label(
            text="Excel no seleccionado",
            size_hint_y=None,
            height=dp(38),
            halign="left",
            valign="middle",
        )
        self.status_label.bind(size=lambda instance, value: setattr(instance, "text_size", value))
        self.root_layout.add_widget(self.status_label)

        top_bar = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(8))

        choose_btn = Button(text="Elegir Excel", size_hint_x=None, width=dp(130))
        choose_btn.bind(on_release=self.choose_excel)
        top_bar.add_widget(choose_btn)

        self.search_input = TextInput(
            hint_text="Buscar modelo, código, EAN o descripción",
            multiline=False,
            write_tab=False,
        )
        self.search_input.bind(text=self.on_search_text)
        top_bar.add_widget(self.search_input)

        clear_btn = Button(text="Limpiar", size_hint_x=None, width=dp(95))
        clear_btn.bind(on_release=lambda instance: self.clear_search())
        top_bar.add_widget(clear_btn)

        self.root_layout.add_widget(top_bar)

        margin_bar = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
        margin_bar.add_widget(Label(text="Margen:", size_hint_x=None, width=dp(70)))

        for margin in MARGIN_OPTIONS:
            btn = Button(text=format_percent(margin), size_hint_x=None, width=dp(78))
            btn.bind(on_release=lambda instance, m=margin: self.set_default_margin(m))
            margin_bar.add_widget(btn)

        self.root_layout.add_widget(margin_bar)

        self.scroll = ScrollView(do_scroll_x=True, do_scroll_y=True)
        self.table = GridLayout(cols=len(DISPLAY_COLUMNS), size_hint=(None, None), spacing=dp(1))
        self.table.bind(minimum_height=self.table.setter("height"))
        self.table.bind(minimum_width=self.table.setter("width"))
        self.scroll.add_widget(self.table)
        self.root_layout.add_widget(self.scroll)

        self.load_config()
        self.render_table([])

        return self.root_layout

    def app_config_path(self):
        return os.path.join(self.user_data_dir, CONFIG_FILE)

    def load_config(self):
        try:
            path = self.app_config_path()
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    config = json.load(f)
                self.excel_path = config.get("excel_path", "")
                self.current_margin = config.get("margin", DEFAULT_MARGIN)

                if self.excel_path:
                    self.status_label.text = f"Excel recordado: {self.excel_path}"
        except Exception:
            pass

    def save_config(self):
        try:
            Path(self.user_data_dir).mkdir(parents=True, exist_ok=True)
            with open(self.app_config_path(), "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "excel_path": self.excel_path,
                        "margin": self.current_margin,
                    },
                    f,
                    indent=4,
                )
        except Exception:
            pass

    def choose_excel(self, instance):
        if filechooser is None:
            self.show_message("Error", "No se pudo abrir el selector de archivos.")
            return

        filechooser.open_file(
            title="Selecciona el Excel",
            filters=[("Excel", "*.xlsx")],
            on_selection=self.on_excel_selected,
        )

    def on_excel_selected(self, selection):
        if not selection:
            return

        self.excel_path = selection[0]
        self.save_config()
        self.load_excel(self.excel_path)
        self.search()

    def load_excel_if_needed(self):
        if self.rows:
            return True

        if not self.excel_path:
            self.show_message("Excel", "Primero selecciona un archivo Excel .xlsx")
            return False

        return self.load_excel(self.excel_path)

    def load_excel(self, path):
        try:
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
            sheet = workbook[SHEET_NAME]

            headers = []
            for cell in sheet[HEADER_ROW]:
                raw_header = cell.value if cell.value is not None else ""
                canonical = CANONICAL_COLUMNS.get(normalize_text(raw_header), str(raw_header).strip())
                headers.append(canonical)

            rows = []
            for excel_row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
                item = {}
                empty = True

                for idx, value in enumerate(excel_row):
                    if idx >= len(headers):
                        continue
                    key = headers[idx]
                    item[key] = "" if value is None else value
                    if value not in [None, ""]:
                        empty = False

                if empty:
                    continue

                self.calculate_row(item)
                rows.append(item)

            workbook.close()
            self.rows = rows
            self.status_label.text = f"Excel cargado: {path}"
            return True
        except Exception as e:
            self.show_message("Error", f"No se pudo leer el Excel:\n{e}")
            return False

    def calculate_row(self, item):
        dto_raw = item.get("Dto", "")

        try:
            if not is_numeric_discount(dto_raw):
                raise ValueError("Dto no numérico")

            pvp = to_number(item.get("PVP", 0))
            dto = parse_discount(dto_raw)
            coste = pvp * (1 - dto)
            margen = self.current_margin
            precio = coste / (1 - margen) if margen < 1 else 0
        except Exception:
            coste = "-"
            margen = "-"
            precio = "-"

        item["COSTE"] = coste
        item["MARGEN"] = margen
        item["PRECIO"] = precio

    def set_default_margin(self, margin):
        self.current_margin = margin
        self.save_config()

        for item in self.rows:
            try:
                if item.get("COSTE") in ["", "-"]:
                    continue
                coste = float(item.get("COSTE", 0))
                item["MARGEN"] = margin
                item["PRECIO"] = coste / (1 - margin) if margin < 1 else 0
            except Exception:
                pass

        self.search()

    def on_search_text(self, instance, value):
        self.search()

    def clear_search(self):
        self.search_input.text = ""
        self.search()

    def search(self):
        if not self.load_excel_if_needed():
            return

        text = normalize_text(self.search_input.text)

        if text == "":
            self.result_rows = self.rows
        else:
            result = []
            for item in self.rows:
                haystack = " ".join(normalize_text(item.get(col, "")) for col in SEARCH_COLUMNS)
                if text in haystack:
                    result.append(item)
            self.result_rows = result

        self.render_table(self.result_rows)

    def render_table(self, rows):
        self.table.clear_widgets()
        self.table.cols = len(DISPLAY_COLUMNS)

        widths = {
            "Modelo": 120,
            "Código": 105,
            "EAN": 125,
            "Descripción": 360,
            "Dto": 80,
            "PVP": 100,
            "COSTE": 100,
            "MARGEN": 110,
            "PRECIO": 110,
        }

        for col in DISPLAY_COLUMNS:
            header = Button(
                text=f"[b]{col}[/b]",
                markup=True,
                size_hint_x=None,
                width=dp(widths.get(col, 120)),
                size_hint_y=None,
                height=dp(44),
                background_color=(0.22, 0.22, 0.22, 1),
            )
            self.table.add_widget(header)

        for item in rows:
            for col in DISPLAY_COLUMNS:
                if col == "MARGEN" and item.get(col) not in ["", "-"]:
                    widget = self.make_margin_editor(item, widths.get(col, 110))
                else:
                    value = self.display_value(item, col)
                    widget = Button(
                        text=value,
                        size_hint_x=None,
                        width=dp(widths.get(col, 120)),
                        size_hint_y=None,
                        height=dp(44),
                        halign="left",
                        valign="middle",
                        background_color=(0.20, 0.34, 0.20, 1) if col == "PRECIO" else (0.13, 0.13, 0.13, 1),
                    )
                    widget.bind(on_release=lambda instance, row=item: self.show_detail_popup(row))
                self.table.add_widget(widget)

    def make_margin_editor(self, item, width):
        box = BoxLayout(
            orientation="horizontal",
            size_hint_x=None,
            width=dp(width),
            size_hint_y=None,
            height=dp(44),
            spacing=dp(1),
        )

        minus_btn = Button(text="-", size_hint_x=None, width=dp(28))
        value_input = TextInput(
            text=f"{float(item.get('MARGEN', 0)) * 100:.1f}",
            multiline=False,
            input_filter="float",
            halign="center",
        )
        plus_btn = Button(text="+", size_hint_x=None, width=dp(28))

        def apply_margin(new_value):
            try:
                margin_percent = float(str(new_value).replace(",", "."))
                margin_percent = max(0.0, min(99.0, margin_percent))
                margin = margin_percent / 100
                coste = float(item.get("COSTE", 0))
                item["MARGEN"] = margin
                item["PRECIO"] = coste / (1 - margin) if margin < 1 else 0
                value_input.text = f"{margin_percent:.1f}"
                self.render_table(self.result_rows)
            except Exception:
                pass

        minus_btn.bind(on_release=lambda instance: apply_margin(float(value_input.text or 0) - 0.1))
        plus_btn.bind(on_release=lambda instance: apply_margin(float(value_input.text or 0) + 0.1))
        value_input.bind(on_text_validate=lambda instance: apply_margin(value_input.text))
        value_input.bind(focus=lambda instance, focus: apply_margin(value_input.text) if not focus else None)

        box.add_widget(minus_btn)
        box.add_widget(value_input)
        box.add_widget(plus_btn)
        return box

    def display_value(self, item, col):
        value = item.get(col, "")

        if col == "Dto":
            return format_discount_for_display(value)

        if value in ["", "-"]:
            return str(value)

        if col in ["PVP", "COSTE", "PRECIO"]:
            return format_money(value)

        if col == "MARGEN":
            return format_percent(value)

        return str(value)

    def show_detail_popup(self, item):
        pvp = to_number(item.get("PVP", 0))
        dto_excel = item.get("Dto", "")
        default_discount = parse_discount(dto_excel) if is_numeric_discount(dto_excel) and str(dto_excel).strip() != "" else 0.18

        layout = BoxLayout(orientation="vertical", padding=dp(12), spacing=dp(8))

        title = Label(
            text=f"{item.get('Modelo', '')}   {item.get('Código', '')}\n{item.get('EAN', '')}\n{item.get('Descripción', '')}",
            size_hint_y=None,
            height=dp(90),
            halign="left",
            valign="middle",
        )
        title.bind(size=lambda instance, value: setattr(instance, "text_size", value))
        layout.add_widget(title)

        info = Label(
            text=f"PVP: {format_money(pvp)}     Dto Excel: {format_discount_for_display(dto_excel)}",
            size_hint_y=None,
            height=dp(34),
            halign="left",
            valign="middle",
        )
        info.bind(size=lambda instance, value: setattr(instance, "text_size", value))
        layout.add_widget(info)

        grid = GridLayout(cols=3, size_hint_y=None, height=dp(260), spacing=dp(5))
        grid.add_widget(Label(text=""))
        grid.add_widget(Label(text="Opción 1"))
        grid.add_widget(Label(text="Opción 2"))

        grid.add_widget(Label(text="Descuento"))
        desc_1 = TextInput(text=f"{default_discount * 100:.1f}%", multiline=False, halign="center")
        desc_2 = TextInput(text="33.0%", multiline=False, halign="center")
        grid.add_widget(desc_1)
        grid.add_widget(desc_2)

        grid.add_widget(Label(text="Coste"))
        coste_1 = Label(text="")
        coste_2 = Label(text="")
        grid.add_widget(coste_1)
        grid.add_widget(coste_2)

        grid.add_widget(Label(text="Margen"))
        margen_1 = self.popup_margin_control()
        margen_2 = self.popup_margin_control()
        grid.add_widget(margen_1["box"])
        grid.add_widget(margen_2["box"])

        grid.add_widget(Label(text="Precio"))
        precio_1 = Label(text="")
        precio_2 = Label(text="")
        grid.add_widget(precio_1)
        grid.add_widget(precio_2)

        layout.add_widget(grid)

        def recalc(*args):
            try:
                d1 = parse_discount(desc_1.text)
                m1 = float(margen_1["input"].text or 0) / 100
                c1 = pvp * (1 - d1)
                p1 = c1 / (1 - m1) if m1 < 1 else 0
                coste_1.text = format_money(c1)
                precio_1.text = format_money(p1)
            except Exception:
                coste_1.text = "-"
                precio_1.text = "-"

            try:
                d2 = parse_discount(desc_2.text)
                m2 = float(margen_2["input"].text or 0) / 100
                c2 = pvp * (1 - d2)
                p2 = c2 / (1 - m2) if m2 < 1 else 0
                coste_2.text = format_money(c2)
                precio_2.text = format_money(p2)
            except Exception:
                coste_2.text = "-"
                precio_2.text = "-"

        for widget in [desc_1, desc_2, margen_1["input"], margen_2["input"]]:
            widget.bind(text=lambda instance, value: recalc())

        margen_1["minus"].bind(on_release=lambda instance: self.change_popup_margin(margen_1["input"], -0.1))
        margen_1["plus"].bind(on_release=lambda instance: self.change_popup_margin(margen_1["input"], 0.1))
        margen_2["minus"].bind(on_release=lambda instance: self.change_popup_margin(margen_2["input"], -0.1))
        margen_2["plus"].bind(on_release=lambda instance: self.change_popup_margin(margen_2["input"], 0.1))

        close_btn = Button(text="Cerrar", size_hint_y=None, height=dp(46))
        layout.add_widget(close_btn)

        popup = Popup(title="Detalle del producto", content=layout, size_hint=(0.95, 0.85))
        close_btn.bind(on_release=popup.dismiss)
        recalc()
        popup.open()

    def popup_margin_control(self):
        box = BoxLayout(orientation="horizontal", spacing=dp(2))
        minus = Button(text="-", size_hint_x=None, width=dp(36))
        value = TextInput(text=f"{self.current_margin * 100:.1f}", multiline=False, halign="center", input_filter="float")
        plus = Button(text="+", size_hint_x=None, width=dp(36))
        box.add_widget(minus)
        box.add_widget(value)
        box.add_widget(plus)
        return {"box": box, "minus": minus, "input": value, "plus": plus}

    def change_popup_margin(self, input_widget, delta):
        try:
            value = float(input_widget.text or 0) + delta
            value = max(0.0, min(99.0, value))
            input_widget.text = f"{value:.1f}"
        except Exception:
            input_widget.text = f"{self.current_margin * 100:.1f}"

    def show_message(self, title, message):
        content = BoxLayout(orientation="vertical", padding=dp(12), spacing=dp(12))
        content.add_widget(Label(text=message))
        btn = Button(text="Cerrar", size_hint_y=None, height=dp(45))
        content.add_widget(btn)
        popup = Popup(title=title, content=content, size_hint=(0.85, 0.45))
        btn.bind(on_release=popup.dismiss)
        popup.open()


if __name__ == "__main__":
    PriceSearchApp().run()
